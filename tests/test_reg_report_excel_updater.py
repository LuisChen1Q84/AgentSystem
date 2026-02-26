#!/usr/bin/env python3
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from scripts.reg_report_excel_updater import collect_findings, run_update


class RegReportExcelUpdaterTest(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        self.xlsx = Path(self.tmpdir.name) / "report.xlsx"

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "表1"
        ws1["A1"] = "2024年支付机构自我评估报告附表"
        ws1["A36"] = "表1 2025年财付通支付科技有限公司广西分公司支付业务量统计（续）"
        ws1["A37"] = "表1 2025年财付通支付科技有限公司广西分公司支付业务量统计（续）"
        ws1["A57"] = "120,466.55"
        ws1["B57"] = "8,303,087.33"
        ws1["F61"] = "填报日期："
        ws1["G61"] = "2025年2月18日"

        ws4 = wb.create_sheet("表4")
        ws4["A1"] = "2024年银行卡收单特约商户与受理终端统计"
        ws4["C23"] = "广西"
        ws4["C40"] = "小计"
        ws4["C52"] = "合计"
        ws4["D23"] = "19118"
        ws4["D40"] = "19118"
        ws4["D52"] = "19118"
        ws4["F23"] = "309"
        ws4["F40"] = "309"
        ws4["F52"] = "309"
        ws4["A70"] = "填表日期"
        ws4["B70"] = "2025年2月18日"

        ws5 = wb.create_sheet("表5")
        ws5["A1"] = "2024年银行卡收单业务量统计"
        ws5["C22"] = "120,466.55"
        ws5["D22"] = "8,303,087.33"
        ws5["A59"] = "制表人：谌虹宇 | 复核人：余意 | 填报日期：2026年2月25日"
        ws5["D59"] = "填报日期：2026年2月25日"
        ws5["E59"] = "2026年2月25日"

        wb.save(self.xlsx)

    def tearDown(self):
        self.tmpdir.cleanup()

    def test_run_update_with_merchant_count(self):
        changes, wb = run_update(self.xlsx, "2026年2月25日", merchant_count="16100")
        self.assertGreater(len(changes), 0)

        ws1 = wb["表1"]
        self.assertIn("2025年", ws1["A1"].value)
        self.assertIsNone(ws1["A36"].value)
        self.assertEqual(ws1["A57"].value, "102,781.63")
        self.assertEqual(ws1["B57"].value, "7,242,453.37")
        self.assertEqual(ws1["G61"].value, "2026年2月25日")

        ws4 = wb["表4"]
        self.assertEqual(ws4["D23"].value, "16100")
        self.assertEqual(ws4["D40"].value, "16100")
        self.assertEqual(ws4["D52"].value, "16100")
        self.assertEqual(ws4["F23"].value, "204")
        self.assertEqual(ws4["F40"].value, "204")
        self.assertEqual(ws4["F52"].value, "204")
        self.assertEqual(ws4["B70"].value, "2026年2月25日")

        ws5 = wb["表5"]
        self.assertEqual(ws5["C22"].value, "102,781.63")
        self.assertEqual(ws5["D22"].value, "7,242,453.37")
        self.assertNotIn("填报日期", str(ws5["A59"].value))
        self.assertIsNone(ws5["E59"].value)

    def test_collect_findings_detects_residue_and_mismatch(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "表X"
        ws["A1"] = "填报日期：2025年2月18日"
        ws["B1"] = "120466.55"

        findings = collect_findings(wb, "2026年2月25日", merchant_count=None)
        cats = {f.category for f in findings}
        self.assertIn("date_mismatch", cats)
        self.assertIn("old_value_residue", cats)


if __name__ == "__main__":
    unittest.main()
