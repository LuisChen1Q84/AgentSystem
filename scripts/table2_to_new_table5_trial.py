#!/usr/bin/env python3
"""Trial calculation: derive new table5 fields directly from table2 events."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


START_DATE = datetime(2021, 9, 30)
END_DEC = datetime(2025, 12, 31)
END_NOV = datetime(2025, 11, 30)
END_JAN = datetime(2026, 1, 31)


@dataclass
class Event:
    metric: str
    value: float
    event_date: datetime
    month: str
    sheet: str
    scope: str
    location: str
    merchant_type: str
    program_type: str


def load_events(jsonl_path: Path) -> List[Event]:
    out: List[Event] = []
    with jsonl_path.open() as f:
        for line in f:
            r = json.loads(line)
            p = r["payload"]
            out.append(
                Event(
                    metric=r["metric"],
                    value=float(r["value"]),
                    event_date=datetime.strptime(r["event_time"][:10], "%Y-%m-%d"),
                    month=p.get("month") or "",
                    sheet=p.get("sheet") or "",
                    scope=p.get("scope") or "",
                    location=p.get("location") or "",
                    merchant_type=p.get("merchant_type") or "",
                    program_type=p.get("program_type") or "",
                )
            )
    return out


def is_valid_loc(loc: str) -> bool:
    return loc not in ("", "NULL", "\\N")


def is_valid_mt(mt: str) -> bool:
    return mt not in ("", "NULL", "\\N")


def sum_benefit(events: List[Event], end_date: datetime, for_province: bool) -> Dict[str, float]:
    """Return benefit cumulative (万元), keyed by location; plus '_NATIONAL_' key."""
    out = defaultdict(float)
    national = 0.0
    for e in events:
        if e.metric != "benefit_amount_yuan":
            continue
        if not (START_DATE <= e.event_date <= end_date):
            continue
        if e.scope != "national_excluding_listed_city":
            continue
        if not is_valid_mt(e.merchant_type):
            continue
        v_wan = e.value / 10000.0
        national += v_wan
        if for_province:
            if e.merchant_type not in ("小微企业", "个体工商户"):
                continue
            if not is_valid_loc(e.location):
                continue
            out[e.location] += v_wan
    out["_NATIONAL_"] = national
    return out


def sum_merchants_202601(events: List[Event]) -> Tuple[Dict[Tuple[str, str], float], Dict[Tuple[str, str], float]]:
    """Return:
    - nt_map: (location, merchant_type) -> 9折+0费商户数
    - d90_map: (location, merchant_type) -> 9折商户数
    """
    nt_map = defaultdict(float)
    d90_map = defaultdict(float)
    for e in events:
        if e.metric != "merchant_count" or e.month != "202601":
            continue
        if not is_valid_mt(e.merchant_type):
            continue

        # 2026年1月全国*商户数 两张表（scope=national）给的是省份口径当月值
        if e.scope == "national":
            nt_map[(e.location, e.merchant_type)] += e.value
        # 9折专用表可直接拿仅9折口径
        if e.sheet == "2026年1月全国9折费率让利商户数":
            d90_map[(e.location, e.merchant_type)] += e.value
    return nt_map, d90_map


def merchant_pass_202601(events: List[Event]) -> float:
    val = 0.0
    for e in events:
        if e.metric == "registered_merchant_cum" and e.month == "202601":
            if e.merchant_type == "有交易行为的个人":
                val = max(val, e.value)
    return val


def evaluate_against_template(template_path: Path, events: List[Event]) -> Dict[str, object]:
    ws = load_workbook(template_path, data_only=True)[load_workbook(template_path, data_only=True).sheetnames[0]]

    b_dec = sum_benefit(events, END_DEC, for_province=True)
    c_jan = sum_benefit(events, END_JAN, for_province=True)
    c_nov = sum_benefit(events, END_NOV, for_province=True)

    nt_map, d90_map = sum_merchants_202601(events)
    mpass = merchant_pass_202601(events)

    # province rows 3..33 + national row34
    checks = []
    failed = []

    for r in range(3, 35):
        name = str(ws[f"A{r}"].value or "").strip()
        if not name:
            continue

        key = name if r < 34 else "_NATIONAL_"
        # B/C/D
        calc_b = b_dec.get(key, 0.0)
        calc_c = c_jan.get(key, 0.0)
        calc_d = 2 * b_dec.get(key, 0.0) - c_nov.get(key, 0.0)
        for col, calc in (("B", calc_b), ("C", calc_c), ("D", calc_d)):
            tar = float(ws[f"{col}{r}"].value)
            ok = abs(calc - tar) <= 1e-6
            checks.append((r, col, ok))
            if not ok:
                failed.append((r, col, tar, calc))

        # N/T/AD
        if r < 34:
            n_calc = nt_map.get((name, "小微企业"), 0.0)
            t_calc = nt_map.get((name, "个体工商户"), 0.0)
            ad_calc = d90_map.get((name, "小微企业"), 0.0) + d90_map.get((name, "个体工商户"), 0.0)
        else:
            # national row uses scope=national totals from same sheets
            # 全国行需包含 location='NULL' 这部分聚合桶
            n_calc = sum(v for (loc, mt), v in nt_map.items() if mt == "小微企业")
            t_calc = sum(v for (loc, mt), v in nt_map.items() if mt == "个体工商户")
            ad_calc = sum(v for (loc, mt), v in d90_map.items() if mt in ("小微企业", "个体工商户"))

        for col, calc in (("N", n_calc), ("T", t_calc), ("AD", ad_calc)):
            tar = float(ws[f"{col}{r}"].value)
            ok = abs(calc - tar) <= 1e-6
            checks.append((r, col, ok))
            if not ok:
                failed.append((r, col, tar, calc))

        # H derivable from N/T and Y(only national)
        if r < 34:
            h_calc = n_calc + t_calc
        else:
            y_calc = sum(v for (loc, mt), v in d90_map.items() if mt == "有交易行为的个人") + mpass
            y_tar = float(ws["Y34"].value)
            yok = abs(y_calc - y_tar) <= 1e-6
            checks.append((34, "Y", yok))
            if not yok:
                failed.append((34, "Y", y_tar, y_calc))
            h_calc = n_calc + t_calc + y_calc
        h_tar = float(ws[f"H{r}"].value)
        hok = abs(h_calc - h_tar) <= 1e-6
        checks.append((r, "H", hok))
        if not hok:
            failed.append((r, "H", h_tar, h_calc))

    # row39 D (其他类商户数逻辑)
    d39_calc = sum(v for (loc, mt), v in nt_map.items() if mt == "其他")
    d39_tar = float(ws["D39"].value)
    d39_ok = abs(d39_calc - d39_tar) <= 1e-6
    checks.append((39, "D", d39_ok))
    if not d39_ok:
        failed.append((39, "D", d39_tar, d39_calc))

    return {
        "checks_total": len(checks),
        "checks_failed": len(failed),
        "checks_passed": len(checks) - len(failed),
        "failed_samples": failed[:20],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Trial compute new table5 from table2 events.")
    parser.add_argument("--events-jsonl", required=True)
    parser.add_argument("--template", required=True)
    args = parser.parse_args()

    events = load_events(Path(args.events_jsonl))
    result = evaluate_against_template(Path(args.template), events)

    print(f"checks_total={result['checks_total']}")
    print(f"checks_passed={result['checks_passed']}")
    print(f"checks_failed={result['checks_failed']}")
    for s in result["failed_samples"]:
        r, c, tar, calc = s
        print(f"FAILED r={r} col={c} target={tar} calc={calc}")


if __name__ == "__main__":
    main()
