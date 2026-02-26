#!/usr/bin/env python3
"""Backfill 2025 monthly data from old Table5 layout into new Table5 template layout."""

from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


# old_col -> new_col
COL_MAP: List[Tuple[str, str]] = [
    ("B", "B"),
    ("C", "C"),
    ("D", "D"),
    ("K", "G"),
    ("L", "H"),
    ("M", "I"),
    ("N", "J"),
    ("O", "K"),
    ("Q", "M"),
    ("R", "N"),
    ("S", "O"),
    ("T", "P"),
    ("U", "Q"),
    ("W", "S"),
    ("X", "T"),
    ("Y", "U"),
    ("Z", "V"),
    ("AB", "X"),
    ("AC", "Y"),
    ("AD", "Z"),
    ("AE", "AA"),
    ("AK", "AC"),
    ("AL", "AD"),
    ("AM", "AE"),
    ("AN", "AF"),
    ("AO", "AG"),
]


def as_num(v: object) -> Optional[float]:
    if isinstance(v, (int, float)):
        return float(v)
    return None


def text(v: object) -> str:
    return "" if v is None else str(v)


def check_formula_chain(ws_values, row_start: int = 3, row_end: int = 34) -> int:
    failed = 0
    for r in range(row_start, row_end + 1):
        name = text(ws_values[f"A{r}"].value).strip()
        if not name:
            continue

        def n(c: str) -> Optional[float]:
            return as_num(ws_values[f"{c}{r}"].value)

        checks = []
        b, c, e = n("B"), n("C"), n("E")
        if None not in (b, c, e):
            checks.append(abs(e - ((c - b) + c)) <= 1e-6)
        c, d, f = n("C"), n("D"), n("F")
        if None not in (c, d, f) and c != 0:
            checks.append(abs(f - ((c - d) / c)) <= 1e-9)
        h, i, j, k, l = n("H"), n("I"), n("J"), n("K"), n("L")
        if None not in (h, i, j, k, l):
            checks.append(abs(l - (h + i + j + k)) <= 1e-6)
        nn, o, p, q, rr = n("N"), n("O"), n("P"), n("Q"), n("R")
        if None not in (nn, o, p, q, rr):
            checks.append(abs(rr - (nn + o + p + q)) <= 1e-6)
        t, u, v, w = n("T"), n("U"), n("V"), n("W")
        if None not in (t, u, v, w):
            checks.append(abs(w - (t + u + v)) <= 1e-6)
        y, z, aa, ab = n("Y"), n("Z"), n("AA"), n("AB")
        if None not in (y, z, aa, ab):
            checks.append(abs(ab - (y + z + aa)) <= 1e-6)
        ad, ae, af, ag, ah = n("AD"), n("AE"), n("AF"), n("AG"), n("AH")
        if None not in (ad, ae, af, ag, ah):
            checks.append(abs(ah - (ad + ae + af + ag)) <= 1e-6)

        if not all(checks):
            failed += 1

    d39, c39, e39 = as_num(ws_values["D39"].value), as_num(ws_values["C39"].value), as_num(ws_values["E39"].value)
    if None not in (d39, c39, e39) and abs(e39 - (d39 + c39)) > 1e-6:
        failed += 1
    return failed


def fill_month_sheet(ws_old, ws_new) -> int:
    mismatch = 0
    # rows with provinces + 全国
    for r in range(3, 35):
        ws_new[f"A{r}"].value = ws_old[f"A{r}"].value
        for old_col, new_col in COL_MAP:
            ws_new[f"{new_col}{r}"].value = ws_old[f"{old_col}{r}"].value

    # special row 39
    ws_new["A39"].value = ws_old["A39"].value
    ws_new["B39"].value = ws_old["B39"].value
    ws_new["C39"].value = ws_old["C39"].value
    ws_new["D39"].value = ws_old["D39"].value

    # direct data-mapping verification
    for r in range(3, 35):
        for old_col, new_col in COL_MAP:
            ov = ws_old[f"{old_col}{r}"].value
            nv = ws_new[f"{new_col}{r}"].value
            on, nn = as_num(ov), as_num(nv)
            if on is not None and nn is not None:
                if abs(on - nn) > 1e-12:
                    mismatch += 1
            elif text(ov) != text(nv):
                mismatch += 1

    return mismatch


def write_db(
    db_path: Path,
    wb_values,
    month_sheets: List[str],
    cleanup_old: bool,
    all_passed: bool,
) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(db_path)
    cur = con.cursor()

    # new normalized store
    cur.executescript(
        """
        DROP TABLE IF EXISTS table5_new_monthly_values;
        CREATE TABLE table5_new_monthly_values (
          sheet_name TEXT,
          month_idx INTEGER,
          row_idx INTEGER,
          col_letter TEXT,
          value_num REAL,
          value_text TEXT
        );
        """
    )

    for idx, month in enumerate(month_sheets, start=1):
        ws = wb_values[month]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                col = chr(ord("A") + c - 1) if c <= 26 else "A" + chr(ord("A") + c - 27)
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)):
                    cur.execute(
                        "INSERT INTO table5_new_monthly_values VALUES (?,?,?,?,?,?)",
                        (month, idx, r, col, float(v), str(v)),
                    )
                elif v is not None and str(v).strip() != "":
                    cur.execute(
                        "INSERT INTO table5_new_monthly_values VALUES (?,?,?,?,?,?)",
                        (month, idx, r, col, None, str(v)),
                    )

    # cleanup legacy old-table5 data only when full validation passed
    if cleanup_old and all_passed:
        # legacy monthly normalized table
        cur.execute("DROP TABLE IF EXISTS table5_monthly_values")
        # legacy wide-cell table keeps mixed sources, remove old table5 source only
        cur.execute("DELETE FROM workbook_cells WHERE source_file='表5.xlsx'")

    con.commit()
    con.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Backfill old Table5 monthly data into new template.")
    parser.add_argument("--old", required=True, help="Old Table5 workbook path (12 monthly sheets).")
    parser.add_argument("--template", required=True, help="New template workbook path (single sheet template).")
    parser.add_argument("--out", required=True, help="Output workbook path.")
    parser.add_argument("--db", required=True, help="SQLite db path.")
    parser.add_argument("--cleanup-old-db", action="store_true", help="Delete old table5 data from DB if checks pass.")
    args = parser.parse_args()

    old_path = Path(args.old)
    template_path = Path(args.template)
    out_path = Path(args.out)
    db_path = Path(args.db)

    wb_old = load_workbook(old_path, data_only=False)
    month_sheets = [f"2025年{m}月" for m in range(1, 13)]
    for m in month_sheets:
        if m not in wb_old.sheetnames:
            raise SystemExit(f"旧表5缺少工作表: {m}")

    wb_out = load_workbook(template_path, data_only=False)
    template_name = wb_out.sheetnames[0]
    template_ws = wb_out[template_name]
    template_idx = wb_out.sheetnames.index(template_name)
    all_mismatch = 0

    # first month: reuse template sheet, rename
    ws_jan = wb_out[template_name]
    ws_jan.title = "2025年1月"
    all_mismatch += fill_month_sheet(wb_old["2025年1月"], ws_jan)

    # remaining months: copy template
    for m in month_sheets[1:]:
        ws = wb_out.copy_worksheet(template_ws if m == "2025年2月" else wb_out["2025年1月"])
        # above copy source doesn't matter after filling because we override input cells and keep formulas
        ws.title = m
        all_mismatch += fill_month_sheet(wb_old[m], ws)

    # order sheets by month
    ordered = [wb_out[m] for m in month_sheets]
    wb_out._sheets = ordered

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(out_path)

    # validations from cached values
    wb_values = load_workbook(out_path, data_only=True)
    formula_failed = 0
    for m in month_sheets:
        formula_failed += check_formula_chain(wb_values[m])

    all_passed = (all_mismatch == 0 and formula_failed == 0)
    write_db(db_path, wb_values, month_sheets, args.cleanup_old_db, all_passed)

    print(f"old={old_path}")
    print(f"template={template_path}")
    print(f"out={out_path}")
    print(f"month_sheets={len(month_sheets)}")
    print(f"mapping_mismatch={all_mismatch}")
    print(f"formula_failed={formula_failed}")
    print(f"all_passed={int(all_passed)}")
    print(f"cleanup_old_db={int(args.cleanup_old_db and all_passed)}")


if __name__ == "__main__":
    main()
