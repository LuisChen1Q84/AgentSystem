#!/usr/bin/env python3
"""Generic workbook regression comparator."""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import List

from openpyxl import load_workbook
from report_rule_config import load_rules


def parse_cells(cells_arg: str) -> List[str]:
    return [x.strip() for x in cells_arg.split(",") if x.strip()]


def main() -> None:
    parser = argparse.ArgumentParser(description="Compare workbook cells for regression checks.")
    parser.add_argument("--expected", required=True)
    parser.add_argument("--actual", required=True)
    parser.add_argument("--sheet", default="")
    parser.add_argument("--cells", default="", help="comma-separated cells, e.g. B34,C34,D34")
    parser.add_argument("--tolerance", type=float, default=-1.0)
    parser.add_argument("--rule-section", default="", help="e.g. table4")
    parser.add_argument("--rule-cells-key", default="key_cells")
    parser.add_argument("--rule-tol-key", default="compare_tolerance")
    args = parser.parse_args()

    rules = load_rules()
    tol = args.tolerance
    cells: List[str] = parse_cells(args.cells)
    if args.rule_section:
        sec = rules.get(args.rule_section, {})
        if not cells:
            cells = list(sec.get(args.rule_cells_key, []))
        if tol < 0:
            tol = float(sec.get(args.rule_tol_key, rules.get("global", {}).get("numeric_tolerance", 1e-6)))
    if tol < 0:
        tol = float(rules.get("global", {}).get("numeric_tolerance", 1e-6))

    ewb = load_workbook(Path(args.expected), data_only=True)
    awb = load_workbook(Path(args.actual), data_only=True)
    es = ewb[args.sheet] if args.sheet else ewb[ewb.sheetnames[0]]
    a_sname = args.sheet if args.sheet and args.sheet in awb.sheetnames else awb.sheetnames[0]
    ac = awb[a_sname]

    if not cells:
        raise SystemExit("没有指定cells，且规则区段也未提供key_cells")

    diffs = []
    for cell in cells:
        ev, av = es[cell].value, ac[cell].value
        if isinstance(ev, (int, float)) and isinstance(av, (int, float)):
            if abs(float(ev) - float(av)) > tol:
                diffs.append((cell, ev, av))
        elif str(ev) != str(av):
            diffs.append((cell, ev, av))

    print(f"sheet_expected={es.title}")
    print(f"sheet_actual={ac.title}")
    print(f"cells_checked={len(cells)}")
    print(f"tolerance={tol}")
    print(f"diff_count={len(diffs)}")
    for cell, ev, av in diffs[:30]:
        print(f"DIFF {cell}: expected={ev} actual={av}")


if __name__ == "__main__":
    main()
