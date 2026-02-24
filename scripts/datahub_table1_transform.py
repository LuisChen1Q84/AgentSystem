#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import json
from pathlib import Path

from openpyxl import load_workbook


# 表1列名到标准metric的映射
METRIC_MAP = {
    "1.【全量商户数】": ("merchant_total_count", "count"),
    "1.【全量主体数】": ("subject_total_count", "count"),
    "2.【近6个月活跃商户数】\n（商户收单交易）": ("merchant_active_6m_count", "count"),
    "2.【近6个月活跃主体数】\n（商户收单交易）": ("subject_active_6m_count", "count"),
    "3【商户当月交易金额】\n(分)（商户收单交易）": ("merchant_txn_amount_cent", "cent"),
    "3.【商户当月交易笔数】\n（商户收单交易）": ("merchant_txn_count", "count"),
    "4.【当月新增账户数】": ("new_account_count", "count"),
    "4.【当月新增主体数】": ("new_subject_count", "count"),
    "5.【当月活跃账户数】": ("active_account_count", "count"),
    "5.【当月活跃主体数】": ("active_subject_count", "count"),
}


def month_to_time(v):
    s = str(v).strip()
    if len(s) == 6 and s.isdigit():
        return dt.datetime.strptime(s, "%Y%m").strftime("%Y-%m-01 00:00:00"), s
    try:
        d = dt.datetime.fromisoformat(s)
        return d.strftime("%Y-%m-01 00:00:00"), d.strftime("%Y%m")
    except ValueError:
        return None, s


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def transform(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h is not None}

    required = ["时间", "是否小微", "行程区划码", "省份", "地市"]
    for k in required:
        if k not in idx:
            raise ValueError(f"缺少必要列: {k}")

    # 找到存在的指标列
    metric_cols = []
    for cn, spec in METRIC_MAP.items():
        if cn in idx:
            metric_cols.append((cn, idx[cn], spec[0], spec[1]))

    if not metric_cols:
        raise ValueError("未识别到任何指标列，请检查表头")

    out_rows = []
    for r in range(2, ws.max_row + 1):
        month_raw = ws.cell(r, idx["时间"]).value
        event_time, month_str = month_to_time(month_raw)
        if not event_time:
            continue

        is_micro = str(ws.cell(r, idx["是否小微"]).value or "").strip()
        region_code = str(ws.cell(r, idx["行程区划码"]).value or "").strip()
        province = str(ws.cell(r, idx["省份"]).value or "").strip()
        city = str(ws.cell(r, idx["地市"]).value or "").strip()
        entity_id = f"{region_code}|{is_micro}" if region_code else is_micro

        for metric_cn, cidx, metric, unit in metric_cols:
            val = to_float(ws.cell(r, cidx).value)
            if val is None:
                continue
            payload = {
                "month": month_str,
                "is_micro": is_micro,
                "region_code": region_code,
                "province": province,
                "city": city,
                "metric_cn": metric_cn,
                "unit": unit,
            }
            out_rows.append(
                {
                    "event_time": event_time,
                    "entity_id": entity_id,
                    "metric": metric,
                    "value": int(val) if float(val).is_integer() else val,
                    "payload": payload,
                }
            )

    return out_rows


def write_csv(rows, path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["event_time", "entity_id", "metric", "value", "payload"])
        w.writeheader()
        for r in rows:
            w.writerow(
                {
                    "event_time": r["event_time"],
                    "entity_id": r["entity_id"],
                    "metric": r["metric"],
                    "value": r["value"],
                    "payload": json.dumps(r["payload"], ensure_ascii=False),
                }
            )


def write_jsonl(rows, path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="表1.xlsx路径")
    parser.add_argument("--csv-out", default="私有数据/import/table1_events.csv")
    parser.add_argument("--jsonl-out", default="私有数据/import/table1_events.jsonl")
    args = parser.parse_args()

    rows = transform(Path(args.xlsx))
    write_csv(rows, Path(args.csv_out))
    write_jsonl(rows, Path(args.jsonl_out))
    print(f"table1转换完成: rows={len(rows)}")
    print(f"csv={args.csv_out}")
    print(f"jsonl={args.jsonl_out}")


if __name__ == "__main__":
    main()
