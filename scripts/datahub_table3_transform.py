#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import json
from pathlib import Path

from openpyxl import load_workbook


METRIC_MAP = {
    '绑定终端数量': ('terminal_bound_count', 'count'),
    '当月新增终端数量': ('terminal_new_month_count', 'count'),
    '暂停终端数量': ('terminal_paused_count', 'count'),
    '五要素齐全终端数量': ('terminal_complete_5elem_count', 'count'),
}


def month_to_time(v):
    s = str(v).strip()
    if s.isdigit() and len(s) == 6:
        d = dt.datetime.strptime(s, '%Y%m')
        return d.strftime('%Y-%m-01 00:00:00'), s
    if s.isdigit() and len(s) == 8:
        d = dt.datetime.strptime(s, '%Y%m%d')
        return d.strftime('%Y-%m-%d 00:00:00'), d.strftime('%Y%m')
    return None, None


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(',', '').strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def transform(xlsx_path: Path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h is not None}

    for k in ('日期', '终端名称', '省份'):
        if k not in idx:
            raise ValueError(f'缺少必要列: {k}')

    metric_cols = []
    for cn, (metric, unit) in METRIC_MAP.items():
        if cn in idx:
            metric_cols.append((cn, idx[cn], metric, unit))
    if not metric_cols:
        raise ValueError('未识别到表3指标列')

    out = []
    for r in range(2, ws.max_row + 1):
        event_time, month = month_to_time(ws.cell(r, idx['日期']).value)
        if not event_time:
            continue

        terminal_name = str(ws.cell(r, idx['终端名称']).value or '').strip()
        province = str(ws.cell(r, idx['省份']).value or '').strip()
        entity_id = f"{province}|{terminal_name}" if province else terminal_name

        for metric_cn, cidx, metric, unit in metric_cols:
            val = to_float(ws.cell(r, cidx).value)
            if val is None:
                continue
            payload = {
                'dataset_id': 'table3',
                'sheet': ws.title,
                'month': month,
                'province': province,
                'terminal_name': terminal_name,
                'metric_cn': metric_cn,
                'unit': unit,
            }
            out.append(
                {
                    'dataset_id': 'table3',
                    'event_time': event_time,
                    'entity_id': entity_id,
                    'metric': metric,
                    'value': int(val) if float(val).is_integer() else val,
                    'payload': payload,
                }
            )
    return out


def write_csv(rows, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8', newline='') as f:
        w = csv.DictWriter(f, fieldnames=['dataset_id', 'event_time', 'entity_id', 'metric', 'value', 'payload'])
        w.writeheader()
        for r in rows:
            w.writerow(
                {
                    'dataset_id': r['dataset_id'],
                    'event_time': r['event_time'],
                    'entity_id': r['entity_id'],
                    'metric': r['metric'],
                    'value': r['value'],
                    'payload': json.dumps(r['payload'], ensure_ascii=False),
                }
            )


def write_jsonl(rows, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8') as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + '\n')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--xlsx', required=True)
    parser.add_argument('--csv-out', default='私有数据/import/table3_events.csv')
    parser.add_argument('--jsonl-out', default='私有数据/import/table3_events.jsonl')
    args = parser.parse_args()

    rows = transform(Path(args.xlsx))
    write_csv(rows, Path(args.csv_out))
    write_jsonl(rows, Path(args.jsonl_out))
    print(f'table3转换完成: rows={len(rows)}')
    print(f'csv={args.csv_out}')
    print(f'jsonl={args.jsonl_out}')


if __name__ == '__main__':
    main()
