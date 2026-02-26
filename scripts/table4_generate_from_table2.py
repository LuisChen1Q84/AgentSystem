#!/usr/bin/env python3
"""Generate quarterly Table4 deliverable from table2_events + monthly new-table5 counts."""

from __future__ import annotations

import argparse
import datetime as dt
import sqlite3
from pathlib import Path
from typing import Dict, Optional, Tuple

from openpyxl import load_workbook
from report_rule_config import load_rules


START_DATE = dt.date(2021, 9, 30)
RULES = load_rules()
T4_RULE = RULES.get("table4", {})


def month_label(yyyymm: str) -> str:
    return f"{yyyymm[:4]}年{int(yyyymm[4:])}月"


def quarter_end_month(year: int, quarter: int) -> str:
    return f"{year}{quarter*3:02d}"


def quarter_months(year: int, quarter: int) -> Tuple[str, str, str]:
    end_m = quarter * 3
    return (f"{year}{end_m-2:02d}", f"{year}{end_m-1:02d}", f"{year}{end_m:02d}")


def quarter_end_date(year: int, quarter: int) -> dt.date:
    m = quarter * 3
    if m == 12:
        return dt.date(year, 12, 31)
    first_next = dt.date(year, m + 1, 1)
    return first_next - dt.timedelta(days=1)


def sum_metric(
    cur: sqlite3.Cursor,
    metric: str,
    merchant_type: str,
    date_start: dt.date,
    date_end: dt.date,
    location: Optional[str] = None,
) -> float:
    sql = """
    SELECT SUM(value)
    FROM table2_events
    WHERE metric=?
      AND merchant_type=?
      AND scope='national_excluding_listed_city'
      AND date(event_time) >= date(?)
      AND date(event_time) <= date(?)
    """
    args = [metric, merchant_type, date_start.isoformat(), date_end.isoformat()]
    if location is not None:
        sql += " AND location=?"
        args.append(location)
    v = cur.execute(sql, args).fetchone()[0]
    return float(v or 0.0)


def get_month_value(cur: sqlite3.Cursor, sheet_name: str, row_idx: int, col_letter: str) -> float:
    v = cur.execute(
        """
        SELECT value_num
        FROM table5_new_monthly_values
        WHERE sheet_name=? AND row_idx=? AND col_letter=?
        """,
        (sheet_name, row_idx, col_letter),
    ).fetchone()
    if v and v[0] is not None:
        return float(v[0])

    # Fallback for formula cells that may not have cached numeric values in DB
    if col_letter == "R":
        return (
            get_month_value(cur, sheet_name, row_idx, "N")
            + get_month_value(cur, sheet_name, row_idx, "O")
            + get_month_value(cur, sheet_name, row_idx, "P")
            + get_month_value(cur, sheet_name, row_idx, "Q")
        )
    if col_letter == "W":
        return (
            get_month_value(cur, sheet_name, row_idx, "T")
            + get_month_value(cur, sheet_name, row_idx, "U")
            + get_month_value(cur, sheet_name, row_idx, "V")
        )
    if col_letter == "AB":
        return (
            get_month_value(cur, sheet_name, row_idx, "Y")
            + get_month_value(cur, sheet_name, row_idx, "Z")
            + get_month_value(cur, sheet_name, row_idx, "AA")
        )
    if col_letter == "E" and row_idx == 39:
        return get_month_value(cur, sheet_name, 39, "D") + get_month_value(cur, sheet_name, 39, "C")
    return 0.0


def fill_section_small_or_individual(
    ws,
    cur: sqlite3.Cursor,
    merchant_type: str,
    col_fee: str,
    col_txn_count: str,
    col_benefit_wan: str,
    col_beneficiary: str,
    year: int,
    quarter: int,
    month_sheet: str,
    beneficiary_col_from_monthly: str,
) -> None:
    q_start = dt.date(year, quarter * 3 - 2, 1)
    q_end = quarter_end_date(year, quarter)

    # province rows
    for r in range(5, 36):
        name = ws[f"A{r}"].value if col_fee == "B" else ws[f"G{r}"].value
        if not name:
            continue
        name = str(name).strip()
        if not name:
            continue

        fee = sum_metric(cur, "fee_income_yuan", merchant_type, q_start, q_end, location=name) / 10000.0
        txn = sum_metric(cur, "txn_count", merchant_type, q_start, q_end, location=name)
        benefit = sum_metric(cur, "benefit_amount_yuan", merchant_type, q_start, q_end, location=name) / 10000.0
        bene_cnt = get_month_value(cur, month_sheet, r - 2, beneficiary_col_from_monthly)

        ws[f"{col_fee}{r}"].value = fee
        ws[f"{col_txn_count}{r}"].value = int(round(txn))
        ws[f"{col_benefit_wan}{r}"].value = round(benefit, 5)
        ws[f"{col_beneficiary}{r}"].value = int(round(bene_cnt))

    # national row 36
    fee_n = sum_metric(cur, "fee_income_yuan", merchant_type, q_start, q_end) / 10000.0
    txn_n = sum_metric(cur, "txn_count", merchant_type, q_start, q_end)
    benefit_n = sum_metric(cur, "benefit_amount_yuan", merchant_type, q_start, q_end) / 10000.0
    bene_n = get_month_value(cur, month_sheet, 34, beneficiary_col_from_monthly)
    ws[f"{col_fee}36"].value = fee_n
    ws[f"{col_txn_count}36"].value = int(round(txn_n))
    ws[f"{col_benefit_wan}36"].value = round(benefit_n, 5)
    ws[f"{col_beneficiary}36"].value = int(round(bene_n))

    # cumulative row 38
    fee_c = sum_metric(cur, "fee_income_yuan", merchant_type, START_DATE, q_end) / 10000.0
    txn_c = sum_metric(cur, "txn_count", merchant_type, START_DATE, q_end)
    benefit_c = sum_metric(cur, "benefit_amount_yuan", merchant_type, START_DATE, q_end) / 10000.0
    ws[f"{col_fee}38"].value = fee_c
    ws[f"{col_txn_count}38"].value = int(round(txn_c))
    ws[f"{col_benefit_wan}38"].value = round(benefit_c, 5)


def fill_person_other(
    ws,
    cur: sqlite3.Cursor,
    year: int,
    quarter: int,
    month_sheet: str,
) -> None:
    q_start = dt.date(year, quarter * 3 - 2, 1)
    q_end = quarter_end_date(year, quarter)

    # personal -> row43 B/C/D, row44 B/C/D
    fee_q = sum_metric(cur, "fee_income_yuan", "有交易行为的个人", q_start, q_end) / 10000.0
    txn_q = sum_metric(cur, "txn_count", "有交易行为的个人", q_start, q_end)
    ben_q = sum_metric(cur, "benefit_amount_yuan", "有交易行为的个人", q_start, q_end) / 10000.0
    ws["B43"].value = fee_q
    ws["C43"].value = int(round(txn_q))
    ws["D43"].value = round(ben_q, 5)
    ws["E43"].value = int(round(get_month_value(cur, month_sheet, 34, "AB")))

    fee_c = sum_metric(cur, "fee_income_yuan", "有交易行为的个人", START_DATE, q_end) / 10000.0
    txn_c = sum_metric(cur, "txn_count", "有交易行为的个人", START_DATE, q_end)
    ben_c = sum_metric(cur, "benefit_amount_yuan", "有交易行为的个人", START_DATE, q_end) / 10000.0
    ws["B44"].value = fee_c
    ws["C44"].value = int(round(txn_c))
    ws["D44"].value = round(ben_c, 5)

    # other -> row43 H/I/J, row44 H/I/J
    fee_q_o = sum_metric(cur, "fee_income_yuan", "其他", q_start, q_end) / 10000.0
    txn_q_o = sum_metric(cur, "txn_count", "其他", q_start, q_end)
    ben_q_o = sum_metric(cur, "benefit_amount_yuan", "其他", q_start, q_end) / 10000.0
    ws["H43"].value = fee_q_o
    ws["I43"].value = int(round(txn_q_o))
    ws["J43"].value = round(ben_q_o, 5)
    ws["K43"].value = int(round(get_month_value(cur, month_sheet, 39, "E")))

    fee_c_o = sum_metric(cur, "fee_income_yuan", "其他", START_DATE, q_end) / 10000.0
    txn_c_o = sum_metric(cur, "txn_count", "其他", START_DATE, q_end)
    ben_c_o = sum_metric(cur, "benefit_amount_yuan", "其他", START_DATE, q_end) / 10000.0
    ws["H44"].value = fee_c_o
    ws["I44"].value = int(round(txn_c_o))
    ws["J44"].value = round(ben_c_o, 5)


def compare_key_cells(ws_calc, ws_ref) -> int:
    cells = T4_RULE.get(
        "key_cells",
        [
            "B36", "C36", "D36", "E36", "H36", "I36", "J36", "K36",
            "B38", "C38", "D38", "H38", "I38", "J38",
            "B43", "C43", "D43", "E43", "H43", "I43", "J43", "K43",
            "B44", "C44", "D44", "H44", "I44", "J44",
        ],
    )
    tol = float(T4_RULE.get("compare_tolerance", 1e-4))
    diff = 0
    for c in cells:
        a, b = ws_calc[c].value, ws_ref[c].value
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            if abs(float(a) - float(b)) > tol:
                diff += 1
        elif str(a) != str(b):
            diff += 1
    return diff


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate Table4 quarter sheet from table2 events")
    parser.add_argument("--db", required=True)
    parser.add_argument("--source", required=True, help="source 表4.xlsx (for template styles/layout)")
    parser.add_argument("--year", type=int, required=True)
    parser.add_argument("--quarter", type=int, choices=[1, 2, 3, 4], required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--reference", default="", help="optional reference 表4.xlsx for key-cell compare")
    args = parser.parse_args()

    db = Path(args.db)
    src = Path(args.source)
    out = Path(args.out)
    year, quarter = args.year, args.quarter
    quarter_name = f"{year}Q{quarter}"
    month_sheet = month_label(quarter_end_month(year, quarter))

    con = sqlite3.connect(db)
    cur = con.cursor()
    for t in ("table2_events", "table5_new_monthly_values"):
        if not cur.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name=?", (t,)).fetchone()[0]:
            raise SystemExit(f"数据库缺少表: {t}")

    wb = load_workbook(src)
    if quarter_name not in wb.sheetnames:
        raise SystemExit(f"source缺少sheet: {quarter_name}")
    ws = wb[quarter_name]

    fill_section_small_or_individual(
        ws, cur, "小微企业", "B", "C", "D", "E", year, quarter, month_sheet, "R"
    )
    fill_section_small_or_individual(
        ws, cur, "个体工商户", "H", "I", "J", "K", year, quarter, month_sheet, "W"
    )
    fill_person_other(ws, cur, year, quarter, month_sheet)

    # keep only target quarter sheet for deliverable
    for s in list(wb.sheetnames):
        if s != quarter_name:
            del wb[s]
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    ref_diff = None
    if args.reference:
        ref_wb = load_workbook(args.reference, data_only=True)
        ref_ws = ref_wb[quarter_name]
        calc_wb = load_workbook(out, data_only=True)
        calc_ws = calc_wb[quarter_name]
        ref_diff = compare_key_cells(calc_ws, ref_ws)

    print(f"quarter={quarter_name}")
    print(f"month_sheet={month_sheet}")
    print(f"out={out}")
    if ref_diff is not None:
        print(f"reference_keycell_diff={ref_diff}")
    con.close()


if __name__ == "__main__":
    main()
