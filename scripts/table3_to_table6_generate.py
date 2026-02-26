#!/usr/bin/env python3
"""Generate monthly Table6 from table3_events in sqlite."""

from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path
from typing import Dict, Optional, Tuple

from openpyxl import load_workbook
from report_rule_config import load_rules


RULES = load_rules()
T6_RULE = RULES.get("table6", {})
ROW_RANGE = range(int(T6_RULE.get("rows_start", 2)), int(T6_RULE.get("rows_end", 37)) + 1)
FORMULA_TOL = float(T6_RULE.get("formula_tolerance", 1e-9))
REGION_ALIAS = T6_RULE.get("region_alias", {})


def prev_month(yyyymm: str) -> str:
    y, m = int(yyyymm[:4]), int(yyyymm[4:])
    if m == 1:
        return f"{y-1}12"
    return f"{y}{m-1:02d}"


def month_label(yyyymm: str) -> str:
    return f"{yyyymm[:4]}年{int(yyyymm[4:])}月"


def normalize_region(label: str) -> Tuple[str, ...]:
    x = label.strip()
    if x in REGION_ALIAS:
        return tuple(REGION_ALIAS[x])
    return (x,)


def fetch_t3(
    cur: sqlite3.Cursor,
    month: str,
) -> Dict[Tuple[str, str, str], float]:
    q = """
    SELECT province, terminal_name, metric, SUM(value)
    FROM table3_events
    WHERE month=?
    GROUP BY province, terminal_name, metric
    """
    return {(p, t, m): float(v or 0.0) for p, t, m, v in cur.execute(q, (month,))}


def fetch_prev_table6_values(cur: sqlite3.Cursor, prev_label: str) -> Dict[Tuple[int, str], float]:
    exists = cur.execute(
        "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='table6_monthly_values'"
    ).fetchone()[0]
    if not exists:
        return {}
    q = """
    SELECT row_idx, col_letter, value_num
    FROM table6_monthly_values
    WHERE sheet_name=? AND value_num IS NOT NULL
    """
    return {(int(r), c): float(v) for r, c, v in cur.execute(q, (prev_label,))}


def get_sum(
    data: Dict[Tuple[str, str, str], float],
    provinces: Tuple[str, ...],
    terminal: str,
    metric: str,
) -> float:
    return sum(data.get((p, terminal, metric), 0.0) for p in provinces)


def validate(ws_values) -> int:
    failed = 0
    for r in ROW_RANGE:
        if not ws_values[f"B{r}"].value:
            continue

        def n(col: str) -> Optional[float]:
            v = ws_values[f"{col}{r}"].value
            return float(v) if isinstance(v, (int, float)) else None

        c, d, e = n("C"), n("D"), n("E")
        if None not in (c, d, e) and abs(e - (c + d)) > FORMULA_TOL:
            failed += 1
        e, f, g = n("E"), n("F"), n("G")
        if None not in (e, f, g) and f != 0 and abs(g - ((e - f) / f)) > FORMULA_TOL:
            failed += 1
        h, i, j, k = n("H"), n("I"), n("J"), n("K")
        if None not in (h, i, j, k) and (abs(j - (h + i)) > FORMULA_TOL or abs(k - (j - e)) > FORMULA_TOL):
            failed += 1
        l, m, nn, o, p, q, rr = n("L"), n("M"), n("N"), n("O"), n("P"), n("Q"), n("R")
        if None not in (l, m, nn, o, p, q, rr):
            if abs(nn - (l + m)) > FORMULA_TOL or abs(q - (o + p)) > FORMULA_TOL or abs(rr - (q - nn)) > FORMULA_TOL:
                failed += 1
    return failed


def persist_table6_month(cur: sqlite3.Cursor, ws_values, sheet_name: str) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS table6_monthly_values (
          sheet_name TEXT,
          row_idx INTEGER,
          col_letter TEXT,
          value_num REAL,
          value_text TEXT
        )
        """
    )
    cur.execute("DELETE FROM table6_monthly_values WHERE sheet_name=?", (sheet_name,))

    for r in range(1, ws_values.max_row + 1):
        for c in range(1, ws_values.max_column + 1):
            col = chr(ord("A") + c - 1) if c <= 26 else "A" + chr(ord("A") + c - 27)
            v = ws_values.cell(r, c).value
            if isinstance(v, (int, float)):
                cur.execute(
                    "INSERT INTO table6_monthly_values VALUES (?,?,?,?,?)",
                    (sheet_name, r, col, float(v), str(v)),
                )
            elif v is not None and str(v).strip() != "":
                cur.execute(
                    "INSERT INTO table6_monthly_values VALUES (?,?,?,?,?)",
                    (sheet_name, r, col, None, str(v)),
                )


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate table6 monthly report from table3_events.")
    parser.add_argument("--db", required=True)
    parser.add_argument("--template", required=True, help="表6模板路径")
    parser.add_argument("--target-month", required=True, help="YYYYMM")
    parser.add_argument("--out", required=True)
    parser.add_argument("--write-db", action="store_true")
    parser.add_argument("--reference", default="", help="可选: 对照文件")
    args = parser.parse_args()

    target = args.target_month
    if len(target) != 6 or not target.isdigit():
        raise SystemExit("target-month必须是YYYYMM")

    db = Path(args.db)
    con = sqlite3.connect(db)
    cur = con.cursor()

    if not cur.execute(
        "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='table3_events'"
    ).fetchone()[0]:
        raise SystemExit("数据库缺少table3_events，请先入库表3")

    data = fetch_t3(cur, target)
    target_label = month_label(target)
    prev_label = month_label(prev_month(target))
    prev_vals = fetch_prev_table6_values(cur, prev_label)

    wb = load_workbook(args.template, data_only=False)
    ws = wb[wb.sheetnames[0]]
    ws.title = target_label

    # A列写目标月份
    for r in ROW_RANGE:
        if ws[f"B{r}"].value:
            ws[f"A{r}"].value = int(target)

    for r in ROW_RANGE:
        region = ws[f"B{r}"].value
        if not region:
            continue
        provinces = normalize_region(str(region))

        # 主终端
        c = get_sum(data, provinces, "小绿盒", "terminal_bound_count")
        d = get_sum(data, provinces, "青蛙商业版", "terminal_bound_count")
        h = get_sum(data, provinces, "小绿盒", "terminal_complete_5elem_count")
        i = get_sum(data, provinces, "青蛙商业版", "terminal_complete_5elem_count")

        # 辅助终端
        l = get_sum(data, provinces, "静态码牌", "terminal_bound_count")
        m = get_sum(data, provinces, "青蛙标准版", "terminal_bound_count")
        o = get_sum(data, provinces, "静态码牌", "terminal_complete_5elem_count")
        p = get_sum(data, provinces, "青蛙标准版", "terminal_complete_5elem_count")

        ws[f"C{r}"].value = int(round(c))
        ws[f"D{r}"].value = int(round(d))
        ws[f"H{r}"].value = int(round(h))
        ws[f"I{r}"].value = int(round(i))
        ws[f"L{r}"].value = int(round(l))
        ws[f"M{r}"].value = int(round(m))
        ws[f"O{r}"].value = int(round(o))
        ws[f"P{r}"].value = int(round(p))

        # 上月终端总数承接：F列取上月E列
        if (r, "E") in prev_vals:
            ws[f"F{r}"].value = int(round(prev_vals[(r, "E")]))

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    wbv = load_workbook(out, data_only=True)
    wsv = wbv[target_label]
    failed = validate(wsv)

    ref_diff = None
    if args.reference:
        ref = load_workbook(args.reference, data_only=True)
        rws = ref[ref.sheetnames[0]]
        diff = 0
        for r in ROW_RANGE:
            for c in ("A", "C", "D", "F", "H", "I", "L", "M", "O", "P"):
                rv, vv = rws[f"{c}{r}"].value, wsv[f"{c}{r}"].value
                if isinstance(rv, (int, float)) and isinstance(vv, (int, float)):
                    if abs(float(rv) - float(vv)) > 1e-6:
                        diff += 1
                elif str(rv) != str(vv):
                    diff += 1
        ref_diff = diff

    if args.write_db:
        persist_table6_month(cur, wsv, target_label)
        con.commit()

    print(f"target_month={target}")
    print(f"target_label={target_label}")
    print(f"out={out}")
    print(f"formula_failed={failed}")
    print(f"wrote_db={int(args.write_db)}")
    if ref_diff is not None:
        print(f"reference_diff={ref_diff}")
    con.close()


if __name__ == "__main__":
    main()
