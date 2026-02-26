#!/usr/bin/env python3
"""监管年报附表定向更新器 v2（预览/执行/验收一体）。"""

import argparse
import datetime as dt
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

TARGET_DATE = "2026年2月25日"
YEAR_OLD = "2024年"
YEAR_NEW = "2025年"

METRIC_OLD_NEW = {
    "tx_count": ("120,466.55", "102,781.63"),
    "tx_amount": ("8,303,087.33", "7,242,453.37"),
    "other_terminal": ("309", "204"),
}

ROW_RULES = {
    "table1": {
        57: [METRIC_OLD_NEW["tx_count"], METRIC_OLD_NEW["tx_amount"]],
    },
    "table4": {
        23: [METRIC_OLD_NEW["other_terminal"]],
        40: [METRIC_OLD_NEW["other_terminal"]],
        52: [METRIC_OLD_NEW["other_terminal"]],
    },
    "table5": {
        22: [METRIC_OLD_NEW["tx_count"], METRIC_OLD_NEW["tx_amount"]],
        39: [METRIC_OLD_NEW["tx_count"], METRIC_OLD_NEW["tx_amount"]],
        51: [METRIC_OLD_NEW["tx_count"], METRIC_OLD_NEW["tx_amount"]],
        57: [METRIC_OLD_NEW["tx_count"], METRIC_OLD_NEW["tx_amount"]],
    },
}

OLD_VALUE_TOKENS = {
    "120,466.55",
    "120466.55",
    "8,303,087.33",
    "8303087.33",
    "309",
    "2025年2月18日",
    "2024年",
}


@dataclass
class Change:
    sheet: str
    cell: str
    old: str
    new: str
    reason: str


@dataclass
class Finding:
    category: str
    sheet: str
    cell: str
    current: str
    expected: str
    message: str
    severity: str = "ERROR"


def normalize_str(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return f"{v}"
    return str(v)


def compact_numeric_str(s: str) -> str:
    return s.replace(",", "").strip()


def normalize_date_text(v: object) -> str:
    if isinstance(v, dt.datetime):
        return f"{v.year}年{v.month}月{v.day}日"
    if isinstance(v, dt.date):
        return f"{v.year}年{v.month}月{v.day}日"
    s = normalize_str(v)
    m = re.search(r"20\d{2}年\d{1,2}月\d{1,2}日", s)
    if m:
        return m.group(0)
    return ""


def replace_year_in_titles(ws) -> List[Change]:
    changes: List[Change] = []
    max_row = min(ws.max_row, 12)
    max_col = min(ws.max_column, 20)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            text = normalize_str(cell.value)
            if YEAR_OLD in text:
                new_text = text.replace(YEAR_OLD, YEAR_NEW)
                if new_text != text:
                    changes.append(Change(ws.title, f"{get_column_letter(c)}{r}", text, new_text, "标题年份更新"))
                    cell.value = new_text
    return changes


def replace_report_date(ws, target_date: str) -> List[Change]:
    changes: List[Change] = []
    date_pat = re.compile(r"20\d{2}年\d{1,2}月\d{1,2}日")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=min(ws.max_column, 30)):
        has_report_date_label = any(
            ("填报日期" in normalize_str(c.value)) or ("填表日期" in normalize_str(c.value)) for c in row
        )
        if not has_report_date_label:
            continue

        for c in row:
            raw = c.value
            if raw is None:
                continue
            old_text = normalize_str(raw)

            if isinstance(raw, (dt.datetime, dt.date)):
                if normalize_date_text(raw) != target_date:
                    changes.append(Change(ws.title, c.coordinate, old_text, target_date, "填报日期统一更新"))
                    c.value = target_date
                    c.number_format = "General"
                continue

            if date_pat.search(old_text):
                new_text = date_pat.sub(target_date, old_text)
                if new_text != old_text:
                    changes.append(Change(ws.title, c.coordinate, old_text, new_text, "填报日期统一更新"))
                    c.value = new_text

    return changes


def detect_table_key(ws) -> Optional[str]:
    title = ws.title
    if re.search(r"表\s*1", title):
        return "table1"
    if re.search(r"表\s*4", title):
        return "table4"
    if re.search(r"表\s*5", title):
        return "table5"

    top_text = "\n".join(
        normalize_str(ws.cell(r, c).value)
        for r in range(1, min(8, ws.max_row) + 1)
        for c in range(1, min(8, ws.max_column) + 1)
    )
    if "表1" in top_text:
        return "table1"
    if "表4" in top_text:
        return "table4"
    if "表5" in top_text:
        return "table5"
    return None


def update_row_metric_cells(ws, row_idx: int, pairs: Iterable[Tuple[str, str]], reason: str) -> List[Change]:
    changes: List[Change] = []
    row = ws[row_idx]
    pending = list(pairs)

    for cell in row:
        text = normalize_str(cell.value)
        if not text:
            continue
        text_compact = compact_numeric_str(text)

        for i, (old_v, new_v) in enumerate(list(pending)):
            old_compact = compact_numeric_str(old_v)
            if old_compact in text_compact:
                new_text = text.replace(old_v, new_v)
                if new_text == text and text_compact == old_compact:
                    new_text = new_v
                if new_text != text:
                    changes.append(Change(ws.title, cell.coordinate, text, new_text, reason))
                    cell.value = new_text
                pending.pop(i)
                break

        if not pending:
            break

    return changes


def update_cell_value(ws, row_idx: int, col_idx: int, new_value: str, reason: str) -> List[Change]:
    changes: List[Change] = []
    cell = ws.cell(row_idx, col_idx)
    old = normalize_str(cell.value)
    if old != str(new_value):
        changes.append(Change(ws.title, cell.coordinate, old, str(new_value), reason))
        cell.value = str(new_value)
    return changes


def find_rows_containing_labels(ws, labels: Set[str], max_scan_col: int = 10) -> List[int]:
    rows: List[int] = []
    for r in range(1, ws.max_row + 1):
        row_texts = [normalize_str(ws.cell(r, c).value).strip() for c in range(1, min(ws.max_column, max_scan_col) + 1)]
        if any(t in labels for t in row_texts if t):
            rows.append(r)
    return rows


def find_rows_containing_keywords(ws, keywords: Tuple[str, ...]) -> List[int]:
    rows: List[int] = []
    for r in range(1, ws.max_row + 1):
        joined = " ".join(normalize_str(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 12) + 1))
        if all(k in joined for k in keywords):
            rows.append(r)
    return rows


def apply_table_specific_updates(ws, merchant_count: Optional[str] = None) -> List[Change]:
    changes: List[Change] = []
    table_key = detect_table_key(ws)
    if not table_key:
        return changes

    dynamic_rows: Dict[int, List[Tuple[str, str]]] = {}
    if table_key == "table1":
        for r in find_rows_containing_keywords(ws, ("银行卡收单",)):
            dynamic_rows[r] = [METRIC_OLD_NEW["tx_count"], METRIC_OLD_NEW["tx_amount"]]
        for r in find_rows_containing_labels(ws, {"合计"}):
            dynamic_rows[r] = [METRIC_OLD_NEW["tx_count"], METRIC_OLD_NEW["tx_amount"]]

    if table_key == "table4":
        for r in find_rows_containing_labels(ws, {"广西", "小计", "合计"}):
            dynamic_rows[r] = [METRIC_OLD_NEW["other_terminal"]]
        if merchant_count:
            for r in find_rows_containing_labels(ws, {"广西", "小计", "合计"}):
                if not normalize_str(ws.cell(r, 4).value).strip():
                    continue
                changes.extend(update_cell_value(ws, r, 4, merchant_count, "table4特约商户数量更新"))

    if table_key == "table5":
        for r in find_rows_containing_labels(ws, {"广西", "小计", "合计", "商户类别合计"}):
            dynamic_rows[r] = [METRIC_OLD_NEW["tx_count"], METRIC_OLD_NEW["tx_amount"]]
        # 若D59已有完整填报日期文本，E59视为冗余占位并清空
        d59 = normalize_str(ws["D59"].value)
        a59 = normalize_str(ws["A59"].value)
        if "填报日期" in d59 and "填报日期" in a59:
            new_a59 = re.sub(r"[|｜]?\s*填报日期[:：]?\s*20\d{2}年\d{1,2}月\d{1,2}日", "", a59).rstrip(" |｜")
            if new_a59 != a59:
                changes.append(Change(ws.title, "A59", a59, new_a59, "冗余日期清理"))
                ws["A59"].value = new_a59
        if "填报日期" in d59 and ws["E59"].value is not None:
            changes.append(Change(ws.title, "E59", normalize_str(ws["E59"].value), "", "冗余日期清理"))
            ws["E59"].value = None
            ws["E59"].number_format = "General"

    merged_rules: Dict[int, List[Tuple[str, str]]] = dict(dynamic_rows)
    for r, pairs in ROW_RULES.get(table_key, {}).items():
        if r not in merged_rules:
            merged_rules[r] = pairs

    for row_idx, pairs in merged_rules.items():
        if row_idx <= ws.max_row:
            changes.extend(update_row_metric_cells(ws, row_idx, pairs, reason=f"{table_key}业务数据更新"))

    if table_key == "table1":
        a36 = normalize_str(ws["A36"].value)
        a37 = normalize_str(ws["A37"].value)
        if a36 and a36 == a37 and "（续）" in a36:
            changes.append(Change(ws.title, "A36", a36, "", "重复标题清理"))
            ws["A36"].value = None

    return changes


def run_update(xlsx_path: Path, target_date: str, merchant_count: Optional[str] = None):
    wb = load_workbook(xlsx_path)
    all_changes: List[Change] = []
    for ws in wb.worksheets:
        all_changes.extend(replace_year_in_titles(ws))
        all_changes.extend(apply_table_specific_updates(ws, merchant_count=merchant_count))
        all_changes.extend(replace_report_date(ws, target_date))
    return all_changes, wb


def collect_findings(wb, target_date: str, merchant_count: Optional[str] = None) -> List[Finding]:
    findings: List[Finding] = []

    for ws in wb.worksheets:
        # 旧值残留
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for c in row:
                txt = normalize_str(c.value).strip()
                if not txt:
                    continue
                if txt in OLD_VALUE_TOKENS:
                    findings.append(
                        Finding(
                            category="old_value_residue",
                            sheet=ws.title,
                            cell=c.coordinate,
                            current=txt,
                            expected="新值或已清理",
                            message="检测到旧值残留",
                        )
                    )

        # 日期一致性与冗余
        for r in range(1, ws.max_row + 1):
            cells = [ws.cell(r, c) for c in range(1, min(ws.max_column, 30) + 1)]
            has_label = any(("填报日期" in normalize_str(c.value)) or ("填表日期" in normalize_str(c.value)) for c in cells)
            if not has_label:
                continue

            dated_cells: List[Tuple[str, str]] = []
            for c in cells:
                d = normalize_date_text(c.value)
                if d:
                    dated_cells.append((c.coordinate, d))
                    if d != target_date:
                        findings.append(
                            Finding(
                                category="date_mismatch",
                                sheet=ws.title,
                                cell=c.coordinate,
                                current=d,
                                expected=target_date,
                                message="填报/填表日期未统一",
                            )
                        )

            if len(dated_cells) > 1:
                joined = ", ".join(f"{coord}:{d}" for coord, d in dated_cells)
                findings.append(
                    Finding(
                        category="redundant_date",
                        sheet=ws.title,
                        cell=str(r),
                        current=joined,
                        expected="单一日期位",
                        message="同一行出现多个日期位，建议保留一个",
                        severity="WARN",
                    )
                )

    # 表4特约商户数量一致性
    if merchant_count:
        for ws in wb.worksheets:
            if detect_table_key(ws) != "table4":
                continue
            for r in find_rows_containing_labels(ws, {"广西", "小计", "合计"}):
                v = normalize_str(ws.cell(r, 4).value).strip()
                if v and v != str(merchant_count):
                    findings.append(
                        Finding(
                            category="merchant_count_mismatch",
                            sheet=ws.title,
                            cell=f"D{r}",
                            current=v,
                            expected=str(merchant_count),
                            message="特约商户数量与近六个月活跃商户主体数不一致",
                        )
                    )

    # 表1重复标题
    for ws in wb.worksheets:
        if detect_table_key(ws) != "table1":
            continue
        a36, a37 = normalize_str(ws["A36"].value), normalize_str(ws["A37"].value)
        if a36 and a36 == a37 and "（续）" in a36:
            findings.append(
                Finding(
                    category="duplicate_title",
                    sheet=ws.title,
                    cell="A36",
                    current=a36,
                    expected="空白（保留A37）",
                    message="检测到重复续页标题",
                )
            )

    return findings


def write_json(path: Path, payload: Dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_plan(changes: List[Change], out_path: Path) -> None:
    payload = {
        "generated_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_changes": len(changes),
        "changes": [asdict(c) for c in changes],
    }
    write_json(out_path, payload)


def write_verify(findings: List[Finding], out_path: Path) -> None:
    payload = {
        "generated_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_findings": len(findings),
        "errors": sum(1 for f in findings if f.severity == "ERROR"),
        "warnings": sum(1 for f in findings if f.severity == "WARN"),
        "findings": [asdict(f) for f in findings],
    }
    write_json(out_path, payload)


def write_markdown_report(changes: List[Change], findings: List[Finding], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    lines: List[str] = []
    lines.append(f"# Excel更新与验收报告 | {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")
    lines.append("## 变更摘要")
    lines.append("")
    lines.append(f"- 变更总数: {len(changes)}")
    reason_count: Dict[str, int] = {}
    for c in changes:
        reason_count[c.reason] = reason_count.get(c.reason, 0) + 1
    for reason, count in sorted(reason_count.items(), key=lambda x: x[0]):
        lines.append(f"- {reason}: {count}")

    lines.append("")
    lines.append("## 验收摘要")
    lines.append("")
    errors = [f for f in findings if f.severity == "ERROR"]
    warns = [f for f in findings if f.severity == "WARN"]
    lines.append(f"- ERROR: {len(errors)}")
    lines.append(f"- WARN: {len(warns)}")

    if findings:
        lines.append("")
        lines.append("## 验收明细")
        lines.append("")
        for f in findings:
            lines.append(
                f"- [{f.severity}] {f.category} | {f.sheet}:{f.cell} | 当前={f.current} | 期望={f.expected} | {f.message}"
            )

    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def print_change_summary(changes: List[Change]) -> None:
    print(f"total_changes={len(changes)}")
    by_reason: Dict[str, int] = {}
    for c in changes:
        by_reason[c.reason] = by_reason.get(c.reason, 0) + 1
    for reason, count in sorted(by_reason.items(), key=lambda x: x[0]):
        print(f"- {reason}: {count}")


def print_verify_summary(findings: List[Finding]) -> None:
    print(f"total_findings={len(findings)}")
    errors = sum(1 for f in findings if f.severity == "ERROR")
    warns = sum(1 for f in findings if f.severity == "WARN")
    print(f"- errors: {errors}")
    print(f"- warnings: {warns}")


def main() -> None:
    parser = argparse.ArgumentParser(description="监管年报附表定向更新器 v2")
    parser.add_argument("--xlsx", required=True, help="目标Excel文件路径（在原文件上修改）")
    parser.add_argument("--mode", choices=["plan", "apply", "verify", "run"], default="plan")
    parser.add_argument("--target-date", default=TARGET_DATE)
    parser.add_argument("--merchant-count", default="", help="表4特约商户数量（近六个月活跃商户主体数）")
    parser.add_argument("--plan-out", default="日志/自动化执行日志/excel_update_plan.json")
    parser.add_argument("--verify-out", default="日志/自动化执行日志/excel_update_verify.json")
    parser.add_argument("--report-out", default="日志/自动化执行日志/excel_update_report.md")
    parser.add_argument("--confirm-token", default="", help="apply/run模式下必须传 APPLY")
    parser.add_argument("--verify-strict", action="store_true", help="若存在ERROR级验收问题则返回非0")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"文件不存在: {xlsx_path}")

    merchant_count = args.merchant_count.strip() or None
    changes: List[Change] = []

    if args.mode in {"plan", "apply", "run"}:
        changes, wb = run_update(xlsx_path, args.target_date, merchant_count=merchant_count)
        print_change_summary(changes)
        write_plan(changes, Path(args.plan_out))
        print(f"plan_file={args.plan_out}")

        if args.mode in {"apply", "run"}:
            if args.confirm_token != "APPLY":
                raise SystemExit("apply/run模式缺少 --confirm-token APPLY，已阻止写入")
            wb.save(xlsx_path)
            print(f"updated_file={xlsx_path}")

    if args.mode == "verify":
        wb_verify = load_workbook(xlsx_path)
    elif args.mode == "plan":
        wb_verify = wb
    else:
        wb_verify = load_workbook(xlsx_path)

    findings = collect_findings(wb_verify, args.target_date, merchant_count=merchant_count)
    print_verify_summary(findings)
    write_verify(findings, Path(args.verify_out))
    write_markdown_report(changes, findings, Path(args.report_out))
    print(f"verify_file={args.verify_out}")
    print(f"report_file={args.report_out}")

    if args.verify_strict and any(f.severity == "ERROR" for f in findings):
        raise SystemExit(3)


if __name__ == "__main__":
    main()
