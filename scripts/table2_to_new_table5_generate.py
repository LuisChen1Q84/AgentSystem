#!/usr/bin/env python3
"""Generate deliverable monthly Table5 (new template) from table2_events in sqlite."""

from __future__ import annotations

import argparse
import datetime as dt
import sqlite3
from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple

from openpyxl import load_workbook
from report_rule_config import load_rules


RULES = load_rules()
T5_RULE = RULES.get("table5", {})
PROVINCE_ROWS = range(int(T5_RULE.get("province_rows_start", 3)), int(T5_RULE.get("province_rows_end", 33)) + 1)
NATIONAL_ROW = int(T5_RULE.get("national_row", 34))
SPECIAL_ROW = int(T5_RULE.get("special_row", 39))
CARRY_COLS = tuple(T5_RULE.get("carry_cols", []))
FORMULA_TOL = float(T5_RULE.get("formula_tolerance", 1e-9))


def month_label(yyyymm: str) -> str:
    return f"{yyyymm[:4]}年{int(yyyymm[4:])}月"


def prev_month(yyyymm: str) -> str:
    y, m = int(yyyymm[:4]), int(yyyymm[4:])
    if m == 1:
        return f"{y-1}12"
    return f"{y}{m-1:02d}"


def month_end_date(yyyymm: str) -> dt.date:
    y, m = int(yyyymm[:4]), int(yyyymm[4:])
    if m == 12:
        return dt.date(y, 12, 31)
    first_next = dt.date(y, m + 1, 1)
    return first_next - dt.timedelta(days=1)


def row_name_map(ws) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        name = ws[f"A{r}"].value
        if name is not None and str(name).strip():
            out[str(name).strip()] = r
    return out


def fetch_province_benefit_cum(
    cur: sqlite3.Cursor,
    end_date: dt.date,
) -> Dict[str, float]:
    q = """
    SELECT location, SUM(value)/10000.0
    FROM table2_events
    WHERE metric='benefit_amount_yuan'
      AND scope='national_excluding_listed_city'
      AND merchant_type IN ('小微企业','个体工商户')
      AND date(event_time) >= date('2021-09-30')
      AND date(event_time) <= date(?)
      AND location NOT IN ('', 'NULL', '\\N')
    GROUP BY location
    """
    return {loc: float(v or 0.0) for loc, v in cur.execute(q, (end_date.isoformat(),))}


def fetch_national_benefit_cum(cur: sqlite3.Cursor, end_date: dt.date) -> float:
    q = """
    SELECT SUM(value)/10000.0
    FROM table2_events
    WHERE metric='benefit_amount_yuan'
      AND scope='national_excluding_listed_city'
      AND merchant_type NOT IN ('', 'NULL', '\\N')
      AND date(event_time) >= date('2021-09-30')
      AND date(event_time) <= date(?)
    """
    v = cur.execute(q, (end_date.isoformat(),)).fetchone()[0]
    return float(v or 0.0)


def fetch_month_merchant_counts(
    cur: sqlite3.Cursor,
    yyyymm: str,
) -> Dict[Tuple[str, str], float]:
    """9折+0费 count by (location, merchant_type), scope=national."""
    q = """
    SELECT location, merchant_type, SUM(value)
    FROM table2_events
    WHERE metric='merchant_count'
      AND month=?
      AND scope='national'
      AND merchant_type NOT IN ('', 'NULL', '\\N')
    GROUP BY location, merchant_type
    """
    return {(loc, mt): float(v or 0.0) for loc, mt, v in cur.execute(q, (yyyymm,))}


def fetch_month_90_counts(
    cur: sqlite3.Cursor,
    yyyymm: str,
) -> Dict[Tuple[str, str], float]:
    """Only 9折 count by (location, merchant_type), scope=national."""
    q = """
    SELECT location, merchant_type, SUM(value)
    FROM table2_events
    WHERE metric='merchant_count'
      AND month=?
      AND scope='national'
      AND program_type='discount_90'
      AND merchant_type NOT IN ('', 'NULL', '\\N')
    GROUP BY location, merchant_type
    """
    return {(loc, mt): float(v or 0.0) for loc, mt, v in cur.execute(q, (yyyymm,))}


def fetch_merchant_pass(cur: sqlite3.Cursor, yyyymm: str) -> float:
    q = """
    SELECT MAX(value)
    FROM table2_events
    WHERE metric='registered_merchant_cum'
      AND month=?
      AND merchant_type='有交易行为的个人'
    """
    v = cur.execute(q, (yyyymm,)).fetchone()[0]
    return float(v or 0.0)


def fetch_prev_month_values(cur: sqlite3.Cursor, sheet_name: str) -> Dict[Tuple[int, str], float]:
    q = """
    SELECT row_idx, col_letter, value_num
    FROM table5_new_monthly_values
    WHERE sheet_name=? AND value_num IS NOT NULL
    """
    return {(int(r), c): float(v) for r, c, v in cur.execute(q, (sheet_name,))}


def validate(ws_values) -> int:
    failed = 0
    for r in range(3, 35):
        if ws_values[f"A{r}"].value in (None, ""):
            continue

        def n(col: str) -> Optional[float]:
            v = ws_values[f"{col}{r}"].value
            return float(v) if isinstance(v, (int, float)) else None

        checks = []
        b, c, e = n("B"), n("C"), n("E")
        if None not in (b, c, e):
            checks.append(abs(e - ((c - b) + c)) <= FORMULA_TOL)
        c, d, f = n("C"), n("D"), n("F")
        if None not in (c, d, f) and c != 0:
            checks.append(abs(f - ((c - d) / c)) <= FORMULA_TOL)
        h, i, j, k, l = n("H"), n("I"), n("J"), n("K"), n("L")
        if None not in (h, i, j, k, l):
            checks.append(abs(l - (h + i + j + k)) <= FORMULA_TOL)
        nn, o, p, q, rr = n("N"), n("O"), n("P"), n("Q"), n("R")
        if None not in (nn, o, p, q, rr):
            checks.append(abs(rr - (nn + o + p + q)) <= FORMULA_TOL)
        t, u, v, w = n("T"), n("U"), n("V"), n("W")
        if None not in (t, u, v, w):
            checks.append(abs(w - (t + u + v)) <= FORMULA_TOL)
        y, z, aa, ab = n("Y"), n("Z"), n("AA"), n("AB")
        if None not in (y, z, aa, ab):
            checks.append(abs(ab - (y + z + aa)) <= FORMULA_TOL)
        ad, ae, af, ag, ah = n("AD"), n("AE"), n("AF"), n("AG"), n("AH")
        if None not in (ad, ae, af, ag, ah):
            checks.append(abs(ah - (ad + ae + af + ag)) <= FORMULA_TOL)

        if not all(checks):
            failed += 1

    d39 = ws_values["D39"].value
    c39 = ws_values["C39"].value
    e39 = ws_values["E39"].value
    if isinstance(d39, (int, float)) and isinstance(c39, (int, float)) and isinstance(e39, (int, float)):
        if abs(float(e39) - (float(d39) + float(c39))) > FORMULA_TOL:
            failed += 1
    return failed


def persist_month_to_db(cur: sqlite3.Cursor, ws_values, sheet_name: str) -> None:
    cur.execute("DELETE FROM table5_new_monthly_values WHERE sheet_name=?", (sheet_name,))
    for r in range(1, ws_values.max_row + 1):
        for c in range(1, ws_values.max_column + 1):
            col = (
                chr(ord("A") + c - 1) if c <= 26 else "A" + chr(ord("A") + (c - 27))
            )  # max AH in this sheet
            v = ws_values.cell(r, c).value
            if isinstance(v, (int, float)):
                cur.execute(
                    "INSERT INTO table5_new_monthly_values VALUES (?,?,?,?,?,?)",
                    (sheet_name, None, r, col, float(v), str(v)),
                )
            elif v is not None and str(v).strip() != "":
                cur.execute(
                    "INSERT INTO table5_new_monthly_values VALUES (?,?,?,?,?,?)",
                    (sheet_name, None, r, col, None, str(v)),
                )


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate monthly new-table5 from table2_events.")
    parser.add_argument("--db", required=True, help="sqlite path containing table2_events")
    parser.add_argument("--template", required=True, help="new table5 template path")
    parser.add_argument("--target-month", required=True, help="YYYYMM, e.g. 202601")
    parser.add_argument("--out", required=True, help="output xlsx path")
    parser.add_argument("--write-db", action="store_true", help="persist generated month into table5_new_monthly_values")
    parser.add_argument("--reference", default="", help="optional xlsx to compare key fields")
    args = parser.parse_args()

    db_path = Path(args.db)
    template_path = Path(args.template)
    out_path = Path(args.out)
    target = args.target_month
    if len(target) != 6 or not target.isdigit():
        raise SystemExit("target-month必须是YYYYMM")

    prev1 = prev_month(target)
    prev2 = prev_month(prev1)
    target_label = month_label(target)
    prev_label = month_label(prev1)

    con = sqlite3.connect(db_path)
    cur = con.cursor()

    for t in ("table2_events", "table5_new_monthly_values"):
        exists = cur.execute(
            "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name=?", (t,)
        ).fetchone()[0]
        if not exists:
            raise SystemExit(f"数据库缺少必要表: {t}")

    prev_values = fetch_prev_month_values(cur, prev_label)
    if not prev_values:
        raise SystemExit(f"数据库里找不到上月({prev_label})数据，无法填报G/M/S/X/AC和承接列")

    p_benefit_prev1 = fetch_province_benefit_cum(cur, month_end_date(prev1))
    p_benefit_target = fetch_province_benefit_cum(cur, month_end_date(target))
    p_benefit_prev2 = fetch_province_benefit_cum(cur, month_end_date(prev2))

    n_benefit_prev1 = fetch_national_benefit_cum(cur, month_end_date(prev1))
    n_benefit_target = fetch_national_benefit_cum(cur, month_end_date(target))
    n_benefit_prev2 = fetch_national_benefit_cum(cur, month_end_date(prev2))

    nt_counts = fetch_month_merchant_counts(cur, target)
    d90_counts = fetch_month_90_counts(cur, target)
    mpass = fetch_merchant_pass(cur, target)

    wb = load_workbook(template_path, data_only=False)
    ws = wb[wb.sheetnames[0]]
    ws.title = target_label
    name_to_row = row_name_map(ws)

    # 省份行
    for r in PROVINCE_ROWS:
        name = str(ws[f"A{r}"].value or "").strip()
        if not name:
            continue

        # amounts
        b = p_benefit_prev1.get(name, 0.0)
        c = p_benefit_target.get(name, 0.0)
        d = 2 * b - p_benefit_prev2.get(name, 0.0)
        ws[f"B{r}"].value = b
        ws[f"C{r}"].value = c
        ws[f"D{r}"].value = d

        # counts current month from table2
        n = nt_counts.get((name, "小微企业"), 0.0)
        t = nt_counts.get((name, "个体工商户"), 0.0)
        h = n + t
        ad = d90_counts.get((name, "小微企业"), 0.0) + d90_counts.get((name, "个体工商户"), 0.0)
        ws[f"H{r}"].value = int(round(h))
        ws[f"N{r}"].value = int(round(n))
        ws[f"T{r}"].value = int(round(t))
        ws[f"AD{r}"].value = int(round(ad))

        # previous month carried metrics
        ws[f"G{r}"].value = int(round(prev_values.get((r, "H"), 0.0)))
        ws[f"M{r}"].value = int(round(prev_values.get((r, "N"), 0.0)))
        ws[f"S{r}"].value = int(round(prev_values.get((r, "T"), 0.0)))
        ws[f"X{r}"].value = int(round(prev_values.get((r, "Y"), 0.0)))
        ws[f"AC{r}"].value = int(round(prev_values.get((r, "AD"), 0.0)))

        for ccol in CARRY_COLS:
            if (r, ccol) in prev_values:
                ws[f"{ccol}{r}"].value = int(round(prev_values[(r, ccol)]))

    # 全国行
    r = NATIONAL_ROW
    n = sum(v for (loc, mt), v in nt_counts.items() if mt == "小微企业")
    t = sum(v for (loc, mt), v in nt_counts.items() if mt == "个体工商户")
    y = sum(v for (loc, mt), v in d90_counts.items() if mt == "有交易行为的个人") + mpass
    h = n + t + y
    ad = sum(v for (loc, mt), v in d90_counts.items() if mt in ("小微企业", "个体工商户"))
    ws[f"B{r}"].value = n_benefit_prev1
    ws[f"C{r}"].value = n_benefit_target
    ws[f"D{r}"].value = 2 * n_benefit_prev1 - n_benefit_prev2
    ws[f"H{r}"].value = int(round(h))
    ws[f"N{r}"].value = int(round(n))
    ws[f"T{r}"].value = int(round(t))
    ws[f"Y{r}"].value = int(round(y))
    ws[f"AD{r}"].value = int(round(ad))

    ws[f"G{r}"].value = int(round(prev_values.get((r, "H"), 0.0)))
    ws[f"M{r}"].value = int(round(prev_values.get((r, "N"), 0.0)))
    ws[f"S{r}"].value = int(round(prev_values.get((r, "T"), 0.0)))
    ws[f"X{r}"].value = int(round(prev_values.get((r, "Y"), 0.0)))
    ws[f"AC{r}"].value = int(round(prev_values.get((r, "AD"), 0.0)))
    for ccol in CARRY_COLS:
        if (r, ccol) in prev_values:
            ws[f"{ccol}{r}"].value = int(round(prev_values[(r, ccol)]))

    # row39 其他类
    if SPECIAL_ROW <= ws.max_row:
        ws[f"D{SPECIAL_ROW}"].value = int(
            round(sum(v for (loc, mt), v in nt_counts.items() if mt == "其他"))
        )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    wb_vals = load_workbook(out_path, data_only=True)
    ws_vals = wb_vals[target_label]
    failed = validate(ws_vals)

    ref_diff = None
    if args.reference:
        ref_ws = load_workbook(args.reference, data_only=True)
        ref = ref_ws[ref_ws.sheetnames[0]]
        diff = 0
        for rr in list(PROVINCE_ROWS) + [NATIONAL_ROW, SPECIAL_ROW]:
            for col in ("B", "C", "D", "G", "H", "N", "T", "Y", "AD"):
                rv = ref[f"{col}{rr}"].value
                vv = ws_vals[f"{col}{rr}"].value
                if isinstance(rv, (int, float)) and isinstance(vv, (int, float)):
                    if abs(float(rv) - float(vv)) > 1e-6:
                        diff += 1
                elif str(rv) != str(vv):
                    diff += 1
        ref_diff = diff

    if args.write_db:
        persist_month_to_db(cur, ws_vals, target_label)
        con.commit()

    print(f"target_month={target}")
    print(f"target_label={target_label}")
    print(f"prev_label={prev_label}")
    print(f"out={out_path}")
    print(f"formula_failed={failed}")
    print(f"wrote_db={int(args.write_db)}")
    if ref_diff is not None:
        print(f"reference_diff={ref_diff}")
    con.close()


if __name__ == "__main__":
    main()
