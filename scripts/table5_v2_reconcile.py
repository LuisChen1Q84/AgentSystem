#!/usr/bin/env python3
"""Reconcile old Table5 output against the new authoritative template.

What this script does:
1) Uses the new template as source-of-truth.
2) Compares with the previously generated (wrong) workbook and summarizes diffs.
3) Validates key formula logic row-by-row.
4) Writes all cells/diffs/validation results into sqlite for traceability.
5) Exports a corrected workbook for delivery.
"""

from __future__ import annotations

import argparse
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclass
class DiffItem:
    row_name: str
    row_idx: int
    col: str
    old_value: str
    new_value: str
    diff_type: str


def as_text(v: object) -> str:
    return "" if v is None else str(v)


def as_num(v: object) -> Optional[float]:
    if isinstance(v, (int, float)):
        return float(v)
    return None


def validate_formula_logic(ws_values, row_start: int = 3, row_end: int = 34) -> List[Tuple[str, int, str, int, str]]:
    """Return validation tuples: (check_name, row, col, ok, message)."""
    checks: List[Tuple[str, int, str, int, str]] = []

    for r in range(row_start, row_end + 1):
        name = as_text(ws_values[f"A{r}"].value).strip()
        if not name:
            continue

        def num(col: str) -> Optional[float]:
            return as_num(ws_values[f"{col}{r}"].value)

        # E = (C-B)+C
        b, c, e = num("B"), num("C"), num("E")
        if None not in (b, c, e):
            expected = (c - b) + c
            ok = abs(e - expected) <= 1e-6
            checks.append(("E=(C-B)+C", r, "E", int(ok), f"{name}: expected={expected}, actual={e}"))

        # F = (C-D)/C
        c, d, f = num("C"), num("D"), num("F")
        if None not in (c, d, f) and c != 0:
            expected = (c - d) / c
            ok = abs(f - expected) <= 1e-9
            checks.append(("F=(C-D)/C", r, "F", int(ok), f"{name}: expected={expected}, actual={f}"))

        # L = H+I+J+K
        h, i, j, k, l = num("H"), num("I"), num("J"), num("K"), num("L")
        if None not in (h, i, j, k, l):
            expected = h + i + j + k
            ok = abs(l - expected) <= 1e-6
            checks.append(("L=H+I+J+K", r, "L", int(ok), f"{name}: expected={expected}, actual={l}"))

        # R = N+O+P+Q
        n, o, p, q, rr = num("N"), num("O"), num("P"), num("Q"), num("R")
        if None not in (n, o, p, q, rr):
            expected = n + o + p + q
            ok = abs(rr - expected) <= 1e-6
            checks.append(("R=N+O+P+Q", r, "R", int(ok), f"{name}: expected={expected}, actual={rr}"))

        # W = T+U+V
        t, u, v, w = num("T"), num("U"), num("V"), num("W")
        if None not in (t, u, v, w):
            expected = t + u + v
            ok = abs(w - expected) <= 1e-6
            checks.append(("W=T+U+V", r, "W", int(ok), f"{name}: expected={expected}, actual={w}"))

        # AH = AD+AE+AF+AG
        ad, ae, af, ag, ah = num("AD"), num("AE"), num("AF"), num("AG"), num("AH")
        if None not in (ad, ae, af, ag, ah):
            expected = ad + ae + af + ag
            ok = abs(ah - expected) <= 1e-6
            checks.append(("AH=AD+AE+AF+AG", r, "AH", int(ok), f"{name}: expected={expected}, actual={ah}"))

    # row39 special
    d39, c39, e39 = as_num(ws_values["D39"].value), as_num(ws_values["C39"].value), as_num(ws_values["E39"].value)
    if None not in (d39, c39, e39):
        expected = d39 + c39
        ok = abs(e39 - expected) <= 1e-6
        checks.append(("E39=D39+C39", 39, "E", int(ok), f"全国其他类: expected={expected}, actual={e39}"))

    return checks


def compare_with_old(ws_new_values, ws_old_values, max_col: int = 34) -> List[DiffItem]:
    old_rows: Dict[str, int] = {}
    for r in range(1, ws_old_values.max_row + 1):
        name = as_text(ws_old_values[f"A{r}"].value).strip()
        if name:
            old_rows[name] = r

    diffs: List[DiffItem] = []
    for r in range(3, 40):
        name = as_text(ws_new_values[f"A{r}"].value).strip()
        if not name or name not in old_rows:
            continue
        old_r = old_rows[name]

        for c in range(1, max_col + 1):
            col = get_column_letter(c)
            nv = ws_new_values[f"{col}{r}"].value
            ov = ws_old_values[f"{col}{old_r}"].value

            nn = as_num(nv)
            on = as_num(ov)
            if nn is not None and on is not None:
                if abs(nn - on) > 1e-12:
                    diffs.append(DiffItem(name, r, col, as_text(ov), as_text(nv), "numeric_mismatch"))
                continue
            if as_text(nv) != as_text(ov):
                diffs.append(DiffItem(name, r, col, as_text(ov), as_text(nv), "text_or_structure_mismatch"))
    return diffs


def store_cells(cur, source_tag: str, ws) -> None:
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            v = cell.value
            is_formula = int(isinstance(v, str) and v.startswith("="))
            comment = cell.comment.text if cell.comment else None
            cur.execute(
                """
                INSERT INTO table5_v2_cells
                (source_tag, sheet_name, row_idx, col_idx, col_letter, value_text, is_formula, comment_text)
                VALUES (?,?,?,?,?,?,?,?)
                """,
                (
                    source_tag,
                    ws.title,
                    r,
                    c,
                    get_column_letter(c),
                    as_text(v),
                    is_formula,
                    comment,
                ),
            )


def build_schema(cur) -> None:
    cur.executescript(
        """
        DROP TABLE IF EXISTS table5_v2_cells;
        DROP TABLE IF EXISTS table5_v2_diff;
        DROP TABLE IF EXISTS table5_v2_validation;

        CREATE TABLE table5_v2_cells (
          source_tag TEXT,
          sheet_name TEXT,
          row_idx INTEGER,
          col_idx INTEGER,
          col_letter TEXT,
          value_text TEXT,
          is_formula INTEGER,
          comment_text TEXT
        );

        CREATE TABLE table5_v2_diff (
          row_name TEXT,
          row_idx INTEGER,
          col_letter TEXT,
          old_value TEXT,
          new_value TEXT,
          diff_type TEXT
        );

        CREATE TABLE table5_v2_validation (
          check_name TEXT,
          row_idx INTEGER,
          col_letter TEXT,
          ok INTEGER,
          message TEXT
        );
        """
    )


def detect_schema_change(ws_new, ws_old) -> bool:
    # Row-2 headers are the best indicator for this template migration.
    for c in range(1, min(ws_new.max_column, ws_old.max_column) + 1):
        col = get_column_letter(c)
        if as_text(ws_new[f"{col}2"].value) != as_text(ws_old[f"{col}2"].value):
            return True
    return ws_new.max_column != ws_old.max_column


def main() -> None:
    parser = argparse.ArgumentParser(description="Reconcile new Table5 template with previous output.")
    parser.add_argument("--template", required=True, help="Path to authoritative new template xlsx")
    parser.add_argument("--old", required=True, help="Path to previous generated (wrong) xlsx")
    parser.add_argument("--db", required=True, help="Path to sqlite db")
    parser.add_argument("--out", required=True, help="Path to corrected xlsx output")
    args = parser.parse_args()

    template_path = Path(args.template)
    old_path = Path(args.old)
    db_path = Path(args.db)
    out_path = Path(args.out)

    wb_new_formula = load_workbook(template_path, data_only=False)
    wb_new_values = load_workbook(template_path, data_only=True)
    ws_new_formula = wb_new_formula[wb_new_formula.sheetnames[0]]
    ws_new_values = wb_new_values[wb_new_values.sheetnames[0]]

    wb_old_values = load_workbook(old_path, data_only=True)
    ws_old_values = wb_old_values["2026年1月"] if "2026年1月" in wb_old_values.sheetnames else wb_old_values[wb_old_values.sheetnames[0]]

    diffs = compare_with_old(ws_new_values, ws_old_values)
    checks = validate_formula_logic(ws_new_values)
    schema_changed = detect_schema_change(ws_new_values, ws_old_values)

    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    build_schema(cur)
    store_cells(cur, "new_template", ws_new_formula)
    store_cells(cur, "old_output", ws_old_values)

    for d in diffs:
        cur.execute(
            "INSERT INTO table5_v2_diff VALUES (?,?,?,?,?,?)",
            (d.row_name, d.row_idx, d.col, d.old_value, d.new_value, d.diff_type),
        )
    for check_name, row_idx, col_letter, ok, msg in checks:
        cur.execute(
            "INSERT INTO table5_v2_validation VALUES (?,?,?,?,?)",
            (check_name, row_idx, col_letter, ok, msg),
        )
    conn.commit()

    # Export corrected workbook directly from authoritative template.
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb_new_formula.save(out_path)

    total_diffs = len(diffs)
    numeric_diffs = sum(1 for d in diffs if d.diff_type == "numeric_mismatch")
    struct_diffs = total_diffs - numeric_diffs
    total_checks = len(checks)
    failed_checks = sum(1 for _, _, _, ok, _ in checks if ok == 0)

    print(f"template={template_path}")
    print(f"old={old_path}")
    print(f"out={out_path}")
    print(f"schema_changed={int(schema_changed)}")
    print(f"diff_total={total_diffs}")
    print(f"diff_numeric={numeric_diffs}")
    print(f"diff_struct_or_text={struct_diffs}")
    print(f"formula_checks_total={total_checks}")
    print(f"formula_checks_failed={failed_checks}")
    if total_checks:
        print(f"formula_checks_pass_rate={(total_checks - failed_checks) / total_checks:.4f}")
    conn.close()


if __name__ == "__main__":
    main()
