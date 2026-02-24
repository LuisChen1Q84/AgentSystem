#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import json
from pathlib import Path

from openpyxl import load_workbook


METRIC_MAP = {
    "笔数": ("txn_count", "count"),
    "交易笔数": ("txn_count", "count"),
    "交易金额": ("txn_amount_yuan", "yuan"),
    "费率收入": ("fee_income_yuan", "yuan"),
    "让利金额": ("benefit_amount_yuan", "yuan"),
    "商户数": ("merchant_count", "count"),
    "当日享受商户数": ("merchant_count", "count"),
    "累计注册商家": ("registered_merchant_cum", "count"),
}


def parse_time(date_val, month_val):
    if date_val not in (None, "", "\\N"):
        s = str(date_val).strip()
        if s.isdigit() and len(s) == 8:
            d = dt.datetime.strptime(s, "%Y%m%d")
            return d.strftime("%Y-%m-%d 00:00:00"), d.strftime("%Y%m")
        if s.isdigit() and len(s) == 6:
            d = dt.datetime.strptime(s, "%Y%m")
            return d.strftime("%Y-%m-01 00:00:00"), d.strftime("%Y%m")
    if month_val not in (None, "", "\\N"):
        s = str(month_val).strip()
        if s.isdigit() and len(s) == 6:
            d = dt.datetime.strptime(s, "%Y%m")
            return d.strftime("%Y-%m-01 00:00:00"), d.strftime("%Y%m")
    return None, None


def to_float(v):
    if v in (None, "", "\\N"):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if not s or s == "\\N":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def norm_header(v):
    if v is None:
        return ""
    return str(v).strip()


def infer_program(sheet_name):
    if "9折" in sheet_name:
        return "discount_90"
    if "0费率" in sheet_name:
        return "rate_zero"
    if "商家通" in sheet_name:
        return "merchant_pass"
    return "unknown"


def infer_scope(sheet_name):
    if "不含单列市" in sheet_name:
        return "national_excluding_listed_city"
    if "单列市" in sheet_name:
        return "listed_city"
    if "全国" in sheet_name:
        return "national"
    return "unknown"


def build_records(wb):
    out = []
    sheet_stats = []

    for sname in wb.sheetnames:
        ws = wb[sname]
        headers = [norm_header(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        idx = {h: i + 1 for i, h in enumerate(headers) if h}

        date_col = idx.get("日期")
        month_col = idx.get("月份")
        city_col = idx.get("城市")
        province_col = idx.get("省份")
        merchant_col = idx.get("商户类型") or idx.get("主体类型")
        flag930_col = idx.get("是否>=930")

        metric_cols = []
        for h, col in idx.items():
            if h in METRIC_MAP:
                metric, unit = METRIC_MAP[h]
                metric_cols.append((h, col, metric, unit))

        # 忽略完全无指标的sheet
        if not metric_cols:
            sheet_stats.append((sname, 0))
            continue

        program = infer_program(sname)
        scope = infer_scope(sname)
        produced = 0

        for r in range(2, ws.max_row + 1):
            date_val = ws.cell(r, date_col).value if date_col else None
            month_val = ws.cell(r, month_col).value if month_col else None
            event_time, month = parse_time(date_val, month_val)
            if not event_time:
                continue

            city = str(ws.cell(r, city_col).value or "").strip() if city_col else ""
            province = str(ws.cell(r, province_col).value or "").strip() if province_col else ""
            merchant_type = str(ws.cell(r, merchant_col).value or "").strip() if merchant_col else ""
            flag930 = ws.cell(r, flag930_col).value if flag930_col else None
            location = city or province or "全国"

            entity_id = "|".join([
                location,
                merchant_type or "ALL",
                program,
                scope,
            ])
            grain_key = "|".join([
                month or "",
                location,
                merchant_type or "ALL",
                program,
                scope,
            ])

            for metric_cn, cidx, metric, unit in metric_cols:
                val = to_float(ws.cell(r, cidx).value)
                if val is None:
                    continue
                payload = {
                    "sheet": sname,
                    "month": month,
                    "location": location,
                    "city": city,
                    "province": province,
                    "merchant_type": merchant_type,
                    "program_type": program,
                    "scope": scope,
                    "metric_cn": metric_cn,
                    "unit": unit,
                    "is_ge_930": int(flag930) if isinstance(flag930, (int, float)) else None,
                    "grain_key": grain_key,
                }
                out.append(
                    {
                        "event_time": event_time,
                        "entity_id": entity_id,
                        "metric": metric,
                        "value": int(val) if float(val).is_integer() else val,
                        "payload": payload,
                    }
                )
                produced += 1

        sheet_stats.append((sname, produced))
    return out, sheet_stats


def write_csv(rows, path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["event_time", "entity_id", "metric", "value", "payload"])
        w.writeheader()
        for r in rows:
            w.writerow(
                {
                    "event_time": r["event_time"],
                    "entity_id": r["entity_id"],
                    "metric": r["metric"],
                    "value": r["value"],
                    "payload": json.dumps(r["payload"], ensure_ascii=False),
                }
            )


def write_jsonl(rows, path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--csv-out", default="私有数据/import/table2_events.csv")
    parser.add_argument("--jsonl-out", default="私有数据/import/table2_events.jsonl")
    args = parser.parse_args()

    wb = load_workbook(Path(args.xlsx), data_only=True)
    rows, stats = build_records(wb)
    write_csv(rows, Path(args.csv_out))
    write_jsonl(rows, Path(args.jsonl_out))

    print(f"table2转换完成: rows={len(rows)}")
    print(f"csv={args.csv_out}")
    print(f"jsonl={args.jsonl_out}")
    for name, cnt in stats:
        print(f"sheet={name} rows={cnt}")


if __name__ == "__main__":
    main()
